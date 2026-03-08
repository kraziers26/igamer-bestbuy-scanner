import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

COLOR_HEADER_BG   = "1F4E79"
COLOR_HEADER_NEW  = "375623"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_STRONG_BG   = "E2EFDA"   # Green — strong buy (15%+ cheaper)
COLOR_MARGINAL_BG = "FFF2CC"   # Yellow — marginal buy (5-15% cheaper)
COLOR_NO_MATCH_BG = "F2F2F2"   # Grey — no match or BB more expensive
COLOR_STRONG_TEXT = "375623"   # Dark green
COLOR_MARGINAL_TEXT = "833C00" # Dark orange
COLOR_LINK        = "0563C1"

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def build_report(inventory: list[dict], results: list[dict], scan_params: dict) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BB Restock Opportunities"

    # Build lookup: source_description → best deal (lowest BB price vs your cost)
    result_map = {}
    for r in results:
        key = r.get("source_description", "")
        if not key:
            continue
        if key not in result_map:
            result_map[key] = r
        else:
            # Keep the one with the biggest savings
            if r.get("savings_pct", 0) > result_map[key].get("savings_pct", 0):
                result_map[key] = r

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"iGamer Corp — Best Buy Restock Opportunities  |  {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
    c.font = Font(name="Arial", bold=True, size=13, color=COLOR_HEADER_FONT)
    c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # ── Scan params ───────────────────────────────────────────────────────────
    ws.merge_cells("A2:J2")
    brands = ", ".join(scan_params.get("brands", []))
    mode = scan_params.get("mode", "both").title()
    min_pct = scan_params.get("min_savings_pct", 0)
    min_dollar = scan_params.get("min_savings_dollar", 0)
    dollar_str = f"  |  Min Savings: ${min_dollar}+" if min_dollar else ""
    c = ws["A2"]
    c.value = f"Brands: {brands}  |  Mode: {mode}  |  Min Savings: {min_pct}%+ cheaper than your cost{dollar_str}"
    c.font = Font(name="Arial", size=10, italic=True, color=COLOR_HEADER_FONT)
    c.fill = PatternFill("solid", fgColor="2E75B6")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # ── Legend ────────────────────────────────────────────────────────────────
    ws.merge_cells("A3:J3")
    c = ws["A3"]
    c.value = "🟢 Strong Buy (15%+ savings)     🟡 Good Deal (5–14% savings)     ⬜ No match / BB more expensive"
    c.font = Font(name="Arial", size=9, italic=True)
    c.fill = PatternFill("solid", fgColor="D9E1F2")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 16

    # ── Headers ───────────────────────────────────────────────────────────────
    original_headers = ["QTY", "CATEGORY", "DESCRIPTION", "STATUS", "YOUR COST (FOB Miami)"]
    bb_headers       = ["BB PRICE", "YOU SAVE $", "YOU SAVE %", "MATCH TYPE", "BUY LINK"]

    header_row = 4
    for col, header in enumerate(original_headers + bb_headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FONT)
        is_bb = col > len(original_headers)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_NEW if is_bb else COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 32

    # ── Data rows ─────────────────────────────────────────────────────────────
    matched = 0
    no_match = 0
    total_savings = []

    for row_idx, product in enumerate(inventory, header_row + 1):
        desc = product.get("description", "")
        result = result_map.get(desc)
        has_deal = result is not None
        match_type = result.get("match_type", "") if has_deal else ""

        if has_deal:
            savings_pct = result.get("savings_pct", 0)
            if savings_pct >= 15:
                row_bg = COLOR_STRONG_BG
            else:
                row_bg = COLOR_MARGINAL_BG
        else:
            row_bg = COLOR_NO_MATCH_BG

        fill = PatternFill("solid", fgColor=row_bg)

        # Original 5 columns
        values = [
            product.get("quantity", ""),
            product.get("category", ""),
            desc,
            product.get("stock_status", ""),
            product.get("price", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = BORDER
            cell.font = Font(name="Arial", size=9)
            if col == 3:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col == 5:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # BB columns
        if has_deal:
            matched += 1
            bb_price      = result.get("bb_price", 0)
            savings_dollar = result.get("savings_dollar", 0)
            savings_pct    = result.get("savings_pct", 0)
            url            = result.get("url", "")
            total_savings.append(savings_pct)

            text_color = COLOR_STRONG_TEXT if savings_pct >= 15 else COLOR_MARGINAL_TEXT

            # BB Price
            c = ws.cell(row=row_idx, column=6, value=bb_price)
            c.number_format = '$#,##0.00'
            c.fill = fill; c.border = BORDER
            c.font = Font(name="Arial", size=9, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

            # You Save $
            c = ws.cell(row=row_idx, column=7, value=savings_dollar)
            c.number_format = '$#,##0.00'
            c.fill = fill; c.border = BORDER
            c.font = Font(name="Arial", size=9, bold=True, color=text_color)
            c.alignment = Alignment(horizontal="center", vertical="center")

            # You Save %
            c = ws.cell(row=row_idx, column=8, value=savings_pct / 100)
            c.number_format = '0.0%'
            c.fill = fill; c.border = BORDER
            c.font = Font(name="Arial", size=9, bold=True, color=text_color)
            c.alignment = Alignment(horizontal="center", vertical="center")

            # Match Type
            match_label = "🎯 Exact" if match_type == "exact" else "🔍 Similar"
            c = ws.cell(row=row_idx, column=9, value=match_label)
            c.fill = fill; c.border = BORDER
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")

            # Buy Link
            c = ws.cell(row=row_idx, column=10, value="🛒 Buy on Best Buy")
            if url:
                c.hyperlink = url
                c.font = Font(name="Arial", size=9, color=COLOR_LINK, underline="single")
            else:
                c.font = Font(name="Arial", size=9, color="999999")
            c.fill = fill; c.border = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")

        else:
            no_match += 1
            for col in range(6, 11):
                c = ws.cell(row=row_idx, column=col,
                            value="No Match / BB Higher" if col == 6 else "—")
                c.fill = fill; c.border = BORDER
                c.font = Font(name="Arial", size=9, color="999999", italic=True)
                c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 48

    # ── Summary row ───────────────────────────────────────────────────────────
    summary_row = header_row + len(inventory) + 1
    ws.merge_cells(f"A{summary_row}:E{summary_row}")
    c = ws.cell(row=summary_row, column=1)
    c.value = f"✅ {matched} restock deals found  |  ❌ {no_match} no match  |  📦 {len(inventory)} total scanned"
    c.font = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FONT)
    c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[summary_row].height = 22

    if total_savings:
        ws.merge_cells(f"F{summary_row}:J{summary_row}")
        c = ws.cell(row=summary_row, column=6)
        c.value = f"Best saving: {max(total_savings):.1f}%  |  Avg saving: {sum(total_savings)/len(total_savings):.1f}%"
        c.font = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FONT)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER_NEW)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Column widths ─────────────────────────────────────────────────────────
    for i, width in enumerate([8, 18, 65, 10, 18, 12, 12, 12, 14, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:J{header_row}"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = f"/tmp/BB_Restock_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path
