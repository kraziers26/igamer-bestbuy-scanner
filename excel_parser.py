import openpyxl
import re
import logging

logger = logging.getLogger(__name__)

KNOWN_BRANDS = [
    "Asus", "ASUS", "Dell", "HP", "Lenovo", "Apple", "Acer",
    "MSI", "Microsoft", "Samsung", "Gigabyte", "Toshiba",
    "PlayStation", "Nintendo", "Firman"
]

SECTION_HEADER_KEYWORDS = [
    "computers", "price list", "transit", "fob miami", "consoles",
    "electric generator", "desktop - all", "apple computer"
]

def is_section_header(row) -> bool:
    col0 = str(row[0]).strip() if row[0] is not None else ""
    col4 = row[4]
    if col4 is None and col0 and not col0.replace(".", "").isdigit():
        for kw in SECTION_HEADER_KEYWORDS:
            if kw in col0.lower():
                return True
        if any(brand.lower() in col0.lower() for brand in KNOWN_BRANDS):
            return True
    return False

def extract_brand(description: str) -> str:
    for brand in KNOWN_BRANDS:
        if description.lower().startswith(brand.lower()):
            return brand
    return description.split()[0] if description else ""

def extract_model_number(description: str) -> str:
    tokens = description.split()
    skip_words = {
        "GAMING","LAPTOP","DESKTOP","COMPUTER","ALL","IN","ONE","MINI","MAX","PRO","AIR",
        "OMEN","ROG","TUF","ZENBOOK","VIVOBOOK","IDEAPAD","THINKPAD","LEGION","LOQ",
        "NITRO","PREDATOR","ALIENWARE","AURORA","ENVY","OMNIBOOK","ASPIRE","SWIFT",
        "VECTOR","CYBORG","KATANA","STRIX","ZEMBOOK","2025","HX","DUO","OMNIDESKTOP"
    }
    for token in tokens[1:]:
        clean = token.rstrip(".,;:")
        if clean.upper() in skip_words:
            continue
        if re.search(r'\d', clean) and len(clean) >= 5:
            return clean
    return ""

def extract_specs(description: str) -> dict:
    specs = {"cpu": "", "ram": "", "storage": ""}
    ram_match = re.search(r'(\d+)\s*GB\s+(?:RAM|Memory|Unified Memory)', description, re.IGNORECASE)
    if ram_match:
        specs["ram"] = f"{ram_match.group(1)}GB"
    storage_match = re.search(r'(\d+(?:\.\d+)?)\s*(TB|GB)\s+(?:Solid State Drive|SSD|M\.2|NVMe)', description, re.IGNORECASE)
    if storage_match:
        specs["storage"] = f"{storage_match.group(1)}{storage_match.group(2).upper()}"
    cpu_patterns = [
        r'(Intel®?\s+Core™?\s+(?:Ultra\s+)?\w+[\s-]\w+[\s-]?\w*)',
        r'(AMD\s+Ryzen(?:\s+AI)?\s+\w+\s+\w+)',
        r'(Intel®?\s+Core™?\s+\w+[\s-]\w+)',
        r'(Apple\s+M\d+(?:\s+\w+)?)',
    ]
    for pattern in cpu_patterns:
        match = re.search(pattern, description, re.IGNORECASE)
        if match:
            specs["cpu"] = match.group(1).strip()
            break
    return specs

def build_exact_query(brand: str, model_number: str, description: str) -> str:
    if model_number and len(model_number) >= 5:
        return model_number
    words = description.split()
    meaningful = [w for w in words[:5] if w.upper() not in ["GAMING","LAPTOP","DESKTOP","COMPUTER"]]
    return " ".join(meaningful[:4])

def build_spec_query(brand: str, specs: dict, category: str) -> str:
    parts = [brand]
    if specs.get("cpu"):
        cpu = re.sub(r'(intel®?\s+|amd\s+)', '', specs["cpu"], flags=re.IGNORECASE).strip()
        parts.append(cpu[:25])
    if specs.get("ram"):
        parts.append(specs["ram"])
    if specs.get("storage"):
        parts.append(specs["storage"])
    cat = category.strip().upper()
    if "GAMING" in cat:
        parts.append("gaming laptop" if "DESKTOP" not in cat else "gaming desktop")
    elif "DESKTOP" in cat:
        parts.append("desktop")
    return " ".join(parts).strip()

def parse_excel(file_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["PRICE LIST"] if "PRICE LIST" in wb.sheetnames else wb.active

    products = []
    seen = set()

    for row in ws.iter_rows(values_only=True):
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        if is_section_header(row):
            continue

        qty_raw = row[0]
        category = str(row[1]).strip() if row[1] else ""
        description = str(row[2]).strip() if row[2] else ""
        stock_status = str(row[3]).strip() if row[3] else ""
        price_raw = row[4]

        if not description or price_raw is None:
            continue
        try:
            quantity = int(float(str(qty_raw))) if qty_raw is not None else 0
        except (ValueError, TypeError):
            continue
        try:
            price = float(str(price_raw).replace(",", "").replace("$", "").strip())
        except (ValueError, TypeError):
            continue
        if price <= 0:
            continue

        brand = extract_brand(description)
        model_number = extract_model_number(description)
        specs = extract_specs(description)

        dedup_key = model_number or description[:60]
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        products.append({
            "brand": brand,
            "model_number": model_number,
            "description": description,
            "category": category,
            "price": price,
            "quantity": quantity,
            "stock_status": stock_status.upper(),
            "cpu": specs["cpu"],
            "ram": specs["ram"],
            "storage": specs["storage"],
            "search_query": build_exact_query(brand, model_number, description),
            "spec_query": build_spec_query(brand, specs, category),
        })

    logger.info(f"Parsed {len(products)} unique products from {file_path}")
    return products

if __name__ == "__main__":
    products = parse_excel("/mnt/user-data/uploads/__iGamer_Corp__Dubai__-_February_20_-ORDER_HERE-.xlsx")
    print(f"\nTotal products parsed: {len(products)}\n")
    for p in products[:8]:
        print(f"Brand: {p['brand']} | Model#: {p['model_number']} | Price: ${p['price']} | Qty: {p['quantity']}")
        print(f"  Exact query: {p['search_query']}")
        print(f"  Spec query:  {p['spec_query']}")
        print()
